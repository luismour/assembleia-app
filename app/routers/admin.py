import os
import secrets
import uuid
import json
import io
from datetime import datetime, timedelta
from dotenv import load_dotenv
from fastapi import APIRouter, Depends, HTTPException, Header, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List
from sqlalchemy.orm import Session
from passlib.context import CryptContext
from jose import JWTError, jwt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app import models

load_dotenv()

router = APIRouter(prefix="/api")

SECRET_KEY = os.getenv("SECRET_KEY", "chave-secreta-padrao")
ALGORITHM = "HS256"
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# --- MODELOS ---
class SenhaAdmin(BaseModel):
    usuario: str
    senha: str
class NovoAdminInput(BaseModel):
    usuario: str
    senha: str
class AssembleiaInput(BaseModel):
    titulo: str
class PautaInput(BaseModel):
    titulo: str
    tipo: str = "SIMPLES"
    candidatos: List[str] = []
    max_escolhas: int = 1
class StatusPauta(BaseModel):
    status: str
class GrupoNomesInput(BaseModel):
    numero: str
    nomes: List[str]
class DadosEnvioEmail(BaseModel):
    nome: str
    email: str
    token: str
    id: str

# --- FUNÇÕES ---
def get_password_hash(password):
    if len(password) > 70: password = password[:70]
    return pwd_context.hash(password)

def verificar_senha(plain_password, hashed_password):
    if len(plain_password) > 70: plain_password = plain_password[:70]
    return pwd_context.verify(plain_password, hashed_password)

def criar_token_acesso(data: dict):
    to_encode = data.copy()
    to_encode.update({"exp": datetime.utcnow() + timedelta(minutes=120)})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def verificar_admin(x_admin_token: str = Header(None), token_query: str = Query(None), db: Session = Depends(get_db)):
    token = x_admin_token or token_query
    if not token: raise HTTPException(401, "Token required")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user = payload.get("sub")
        if user == "admin": return "admin"
        db_admin = db.query(models.Admin).filter(models.Admin.usuario == user).first()
        if not db_admin: raise HTTPException(401, "Invalid")
        return user
    except JWTError: raise HTTPException(401, "Invalid")

# --- ROTAS ---

@router.post("/admin/login")
def admin_login(dados: SenhaAdmin, db: Session = Depends(get_db)):
    usuario_input = dados.usuario.strip()
    senha_input = dados.senha.strip()

    if usuario_input == "admin" and senha_input == "nrpesaps":
        adm_db = db.query(models.Admin).filter(models.Admin.usuario == "admin").first()
        novo_hash = get_password_hash("nrpesaps")
        if not adm_db: db.add(models.Admin(usuario="admin", senha_hash=novo_hash))
        else: adm_db.senha_hash = novo_hash
        db.commit()
        return {"token": criar_token_acesso(data={"sub": "admin"})}

    adm = db.query(models.Admin).filter(models.Admin.usuario == usuario_input).first()
    if not adm or not verificar_senha(senha_input, adm.senha_hash):
        raise HTTPException(400, "Usuário ou senha incorretos")
    return {"token": criar_token_acesso(data={"sub": adm.usuario})}

@router.get("/admin/lista-para-email", response_model=List[DadosEnvioEmail])
def lista_emails_bulk(x_admin_token: str = Header(None), token: str = Query(None), db: Session = Depends(get_db)):
    verificar_admin(x_admin_token, token, db)
    users = db.query(models.Usuario).filter(models.Usuario.email != None, models.Usuario.email != "").all()
    resultado = []
    for u in users:
        resultado.append({"nome": u.nome, "email": u.email, "token": u.token, "id": u.id})
    return resultado

@router.get("/admins")
def list_admins(db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    return db.query(models.Admin).all()

@router.post("/admins")
def add_admin(d: NovoAdminInput, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    if db.query(models.Admin).filter(models.Admin.usuario == d.usuario).first(): raise HTTPException(400, "Exists")
    db.add(models.Admin(usuario=d.usuario, senha_hash=get_password_hash(d.senha)))
    db.commit()
    return {"msg": "Ok"}

@router.delete("/admins/{nome}")
def del_admin(nome: str, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    if nome == u: raise HTTPException(400, "Self delete")
    db.query(models.Admin).filter(models.Admin.usuario == nome).delete()
    db.commit()
    return {"msg": "Ok"}

@router.get("/telao-dados")
def get_telao(db: Session = Depends(get_db)):
    asm = db.query(models.Assembleia).filter(models.Assembleia.ativa == True).first()
    nome = asm.titulo if asm else "Escoteiros"
    if not asm: return {"evento": nome, "pauta": None}
    pauta = db.query(models.Pauta).filter(models.Pauta.assembleia_id == asm.id, models.Pauta.status == "ABERTA").first()
    if not pauta: pauta = db.query(models.Pauta).filter(models.Pauta.assembleia_id == asm.id).order_by(models.Pauta.id.desc()).first()
    if not pauta: return {"evento": nome, "pauta": None}
    
    votos = db.query(models.Voto).filter(models.Voto.pauta_id == pauta.id).all()
    candidatos = json.loads(pauta.candidatos_str) if pauta.candidatos_str else []
    
    if pauta.tipo == "SIMPLES":
        cont = {"favor":0, "contra":0, "abstencao":0}
        for v in votos:
            val = json.loads(v.escolha_str)
            if val in cont: cont[val] += 1
    else:
        cont = {c: 0 for c in candidatos}
        for v in votos:
            vals = json.loads(v.escolha_str)
            if isinstance(vals, list):
                for c in vals: 
                    if c in cont: cont[c] += 1
            elif vals in cont: cont[vals] += 1
            
    res = "ANDAMENTO"
    if pauta.status == "ENCERRADA":
        if pauta.tipo == "SIMPLES":
            if cont["favor"] > cont["contra"]: res = "APROVADA"
            elif cont["contra"] >= cont["favor"] and len(votos)>0: res = "REPROVADA"
            else: res = "SEM VOTOS"
        else: res = "ELEIÇÃO CONCLUÍDA"
        
    return {"evento": nome, "pauta": {"titulo": pauta.titulo, "status": pauta.status, "tipo": pauta.tipo, "max_escolhas": pauta.max_escolhas, "total_votos": len(votos), "resultados": cont, "resultado_final": res}}

@router.get("/assembleias")
def get_asms(db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    lista = db.query(models.Assembleia).all()
    ativa = db.query(models.Assembleia).filter(models.Assembleia.ativa == True).first()
    return {"lista": lista, "ativa": ativa.id if ativa else None}

@router.post("/assembleias")
def add_asm(d: AssembleiaInput, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    nova = models.Assembleia(id=str(uuid.uuid4()), titulo=d.titulo)
    db.add(nova)
    if not db.query(models.Assembleia).filter(models.Assembleia.ativa == True).first(): nova.ativa = True
    db.commit()
    return nova

# --- NOVAS ROTAS (EDITAR E EXCLUIR EVENTO) ---
@router.put("/assembleias/{id}")
def edit_asm(id: str, d: AssembleiaInput, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    asm = db.query(models.Assembleia).filter(models.Assembleia.id == id).first()
    if not asm:
        raise HTTPException(404, "Evento não encontrado")
    
    asm.titulo = d.titulo
    db.commit()
    return {"msg": "ok", "titulo": asm.titulo}

@router.delete("/assembleias/{id}")
def delete_asm(id: str, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    asm = db.query(models.Assembleia).filter(models.Assembleia.id == id).first()
    if not asm:
        raise HTTPException(404, "Evento não encontrado")
    
    # 1. Buscar todas as pautas deste evento
    pautas = db.query(models.Pauta).filter(models.Pauta.assembleia_id == id).all()
    
    # 2. Para cada pauta, deletar os votos e depois a pauta
    for p in pautas:
        db.query(models.Voto).filter(models.Voto.pauta_id == p.id).delete()
        db.delete(p)
    
    # 3. Deletar a assembleia
    db.delete(asm)
    db.commit()
    return {"msg": "ok"}
# ---------------------------------------------

@router.post("/assembleias/{id}/ativar")
def set_active_asm(id: str, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    db.query(models.Assembleia).update({models.Assembleia.ativa: False})
    target = db.query(models.Assembleia).filter(models.Assembleia.id == id).first()
    if target: target.ativa = True
    db.commit()
    return {"msg": "Ok"}

@router.post("/grupos")
def add_grupo_massa(d: GrupoNomesInput, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    todos = db.query(models.Usuario).filter(models.Usuario.grupo == d.numero).all()
    prox = 1
    if todos:
        try: prox = max([int(x.id.split('-')[1]) for x in todos]) + 1
        except: prox = len(todos) + 1
    novos = []
    for nome in d.nomes:
        uid = f"{d.numero}-{prox}"
        while True:
            t = secrets.token_hex(3).upper()
            if not db.query(models.Usuario).filter(models.Usuario.token == t).first(): break
        usr = models.Usuario(id=uid, nome=nome.strip(), grupo=d.numero, token=t, checkin=False, cpf="", email="")
        db.add(usr)
        novos.append(usr)
        prox += 1
    db.commit()
    return {"msg": "ok", "delegados": novos}

@router.get("/grupos")
def list_grupos(db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    users = db.query(models.Usuario).all()
    agg = {}
    for user in users:
        if user.grupo not in agg: agg[user.grupo] = []
        agg[user.grupo].append(user)
    lst = [{"numero": k, "quantidade": len(v), "delegados": v} for k,v in agg.items()]
    lst.sort(key=lambda x: int(x["numero"]) if x["numero"].isdigit() else x["numero"])
    return lst

@router.post("/usuarios/{token}/checkin")
def toggle_checkin(token: str, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    usr = db.query(models.Usuario).filter(models.Usuario.token == token).first()
    if usr:
        usr.checkin = not usr.checkin
        db.commit()
        return {"msg": "ok"}
    raise HTTPException(404)

@router.delete("/usuarios/{token}")
def del_user(token: str, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    usr = db.query(models.Usuario).filter(models.Usuario.token == token).first()
    if usr:
        db.query(models.Voto).filter(models.Voto.usuario_id == usr.id).delete()
        db.delete(usr)
        db.commit()
        return {"msg": "ok"}
    raise HTTPException(404)

@router.delete("/grupos/{n}")
def del_grp(n: str, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    usrs = db.query(models.Usuario).filter(models.Usuario.grupo == n).all()
    for usr in usrs:
        db.query(models.Voto).filter(models.Voto.usuario_id == usr.id).delete()
        db.delete(usr)
    db.commit()
    return {"msg": "ok"}

@router.get("/dados-admin")
def admin_data(db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    asm = db.query(models.Assembleia).filter(models.Assembleia.ativa == True).first()
    if not asm: return {"pautas": [], "assembleia": "Nenhuma"}
    total_users = db.query(models.Usuario).count()
    pautas = db.query(models.Pauta).filter(models.Pauta.assembleia_id == asm.id).all()
    pautas.reverse()
    res = []
    users_map = {u.id: u for u in db.query(models.Usuario).all()}
    for p in pautas:
        votos = db.query(models.Voto).filter(models.Voto.pauta_id == p.id).all()
        dets = []
        cands = json.loads(p.candidatos_str) if p.candidatos_str else []
        if p.tipo == "SIMPLES":
            cont = {"favor":0, "contra":0, "abstencao":0}
            for v in votos:
                val = json.loads(v.escolha_str)
                if val in cont: cont[val] += 1
                usr = users_map.get(v.usuario_id)
                dets.append({"credencial": v.usuario_id, "nome": usr.nome if usr else "?", "grupo": usr.grupo if usr else "-", "voto": val})
        else:
            cont = {c: 0 for c in cands}
            for v in votos:
                vals = json.loads(v.escolha_str)
                v_str = ""
                if isinstance(vals, list):
                    for c in vals: 
                        if c in cont: cont[c] += 1
                    v_str = ", ".join(vals)
                else: v_str = str(vals)
                usr = users_map.get(v.usuario_id)
                dets.append({"credencial": v.usuario_id, "nome": usr.nome if usr else "?", "grupo": usr.grupo if usr else "-", "voto": v_str})
            cont = dict(sorted(cont.items(), key=lambda i: i[1], reverse=True))
        final = "ANDAMENTO"
        if p.status == "ENCERRADA":
            if p.tipo == "SIMPLES": final = "APROVADA" if cont["favor"] > cont["contra"] else "REPROVADA"
            else: final = "CONCLUÍDA"
        res.append({"id": p.id, "titulo": p.titulo, "status": p.status, "tipo": p.tipo, "candidatos": cands, "max_escolhas": p.max_escolhas, "total_votos": len(votos), "esperados": total_users, "resultados": cont, "resultado_final": final, "votos_detalhados": dets})
    return {"pautas": res, "assembleia": asm.titulo}

@router.post("/pautas")
def add_pauta(d: PautaInput, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    asm = db.query(models.Assembleia).filter(models.Assembleia.ativa == True).first()
    if not asm: raise HTTPException(400, "Sem assembleia")
    nova = models.Pauta(id=str(uuid.uuid4()), titulo=d.titulo, assembleia_id=asm.id, tipo=d.tipo, max_escolhas=d.max_escolhas, candidatos_str=json.dumps(d.candidatos))
    db.add(nova)
    db.commit()
    return nova

@router.put("/pautas/{id}")
def edit_pauta(id: str, d: PautaInput, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    p = db.query(models.Pauta).filter(models.Pauta.id == id).first()
    if not p: raise HTTPException(404)
    p.titulo = d.titulo
    p.tipo = d.tipo
    p.max_escolhas = d.max_escolhas
    p.candidatos_str = json.dumps(d.candidatos)
    db.commit()
    return {"msg": "ok"}

@router.delete("/pautas/{id}")
def del_pauta(id: str, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    p = db.query(models.Pauta).filter(models.Pauta.id == id).first()
    if p:
        db.query(models.Voto).filter(models.Voto.pauta_id == id).delete()
        db.delete(p)
        db.commit()
        return {"msg": "ok"}
    raise HTTPException(404)

@router.post("/pautas/{id}/status")
def set_status(id: str, d: StatusPauta, db: Session = Depends(get_db), u: str = Depends(verificar_admin)):
    p = db.query(models.Pauta).filter(models.Pauta.id == id).first()
    if not p: raise HTTPException(404)
    if d.status == "ABERTA": db.query(models.Pauta).filter(models.Pauta.assembleia_id == p.assembleia_id, models.Pauta.status == "ABERTA").update({models.Pauta.status: "ENCERRADA"})
    p.status = d.status
    db.commit()
    return {"msg": "ok"}

@router.get("/exportar")
def exportar(x_admin_token: str = Header(None), token: str = Query(None), db: Session = Depends(get_db)):
    verificar_admin(x_admin_token, token, db)
    asm = db.query(models.Assembleia).filter(models.Assembleia.ativa == True).first()
    tn = asm.titulo if asm else "Relatorio"
    
    wb = Workbook()
    hf = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    fill = PatternFill(start_color="002d62", end_color="002d62", fill_type="solid")
    ca = Alignment(horizontal='center', vertical='center')
    bd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # ABA 1: RESUMO
    ws1 = wb.active; ws1.title = "Resumo"
    ws1.merge_cells('A1:E1'); ws1['A1'].value=f"RELATÓRIO: {tn.upper()}"; ws1['A1'].font=Font(size=14, bold=True, color="002d62"); ws1['A1'].alignment=ca
    ws1.append(["Ordem", "Pauta", "Tipo", "Status", "Votos"])
    for c in ws1[2]: c.font=hf; c.fill=fill; c.alignment=ca
    
    pautas = db.query(models.Pauta).filter(models.Pauta.assembleia_id == asm.id).all() if asm else []
    for i, p in enumerate(pautas, 1):
        v_count = db.query(models.Voto).filter(models.Voto.pauta_id == p.id).count()
        ws1.append([i, p.titulo, p.tipo, p.status, v_count])
        for cell in ws1[ws1.max_row]: cell.border=bd; cell.alignment=ca
    
    # ABA 2: DETALHAMENTO DE VOTOS
    ws2 = wb.create_sheet("Detalhamento Votos")
    ws2.append(["Pauta", "ID", "Nome", "Grupo", "Voto"])
    for c in ws2[1]: c.font=hf; c.fill=fill; c.alignment=ca
    
    users_map = {u.id: u for u in db.query(models.Usuario).all()}
    for p in pautas:
        votos = db.query(models.Voto).filter(models.Voto.pauta_id == p.id).all()
        for v in votos:
            usr = users_map.get(v.usuario_id)
            val = json.loads(v.escolha_str)
            if p.tipo == "SIMPLES":
                vf = val.upper()
                if val == "favor": vf = "A FAVOR"
            else:
                vf = ", ".join(val) if isinstance(val, list) else str(val)
            ws2.append([p.titulo, v.usuario_id, usr.nome if usr else "?", f"GE {usr.grupo if usr else '-'}", vf])
            for cell in ws2[ws2.max_row]: cell.border=bd
    ws2.column_dimensions['A'].width=40; ws2.column_dimensions['C'].width=35; ws2.column_dimensions['E'].width=50
    
    ws3 = wb.create_sheet("Credenciados (Emails)")
    ws3.append(["ID", "Nome", "Grupo", "Email", "CPF", "Token", "Check-in"])
    for c in ws3[1]: c.font=hf; c.fill=fill; c.alignment=ca
    
    todos_usuarios = db.query(models.Usuario).all()
    for u in todos_usuarios:
        ws3.append([u.id, u.nome, u.grupo, u.email, u.cpf, u.token, "SIM" if u.checkin else "NÃO"])
        for cell in ws3[ws3.max_row]: cell.border=bd
    
    ws3.column_dimensions['B'].width=35; ws3.column_dimensions['D'].width=35; ws3.column_dimensions['E'].width=15
    
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, headers={'Content-Disposition': f'attachment; filename="{tn}.xlsx"'}, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')